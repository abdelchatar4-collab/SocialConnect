# Copyright (C) 2025 ABDEL KADER CHATAR
# SocialConnect est un logiciel libre : vous pouvez le redistribuer et/ou le modifier selon les termes de la Licence Publique Générale GNU telle que publiée par la Free Software Foundation, soit la version 3 de la licence, soit (à votre convenance) toute version ultérieure.
#
# Ce programme est distribué dans l'espoir qu'il sera utile, mais SANS AUCUNE GARANTIE ; sans même la garantie implicite de COMMERCIALISATION ou d'ADÉQUATION À UN USAGE PARTICULIER. Voir la Licence Publique Générale GNU pour plus de détails.

import sys
import json
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment # Ajout de Alignment
from openpyxl.utils import get_column_letter
# dateutil.parser sera importé dans la fonction format_date pour gérer une potentielle ImportError

def format_date(date_str):
    """Formats a date string to DD/MM/YYYY (European format) with robust fallback."""
    if not date_str:
        return ""

    # Si c'est déjà une chaîne au format DD/MM/YYYY, la retourner telle quelle
    if isinstance(date_str, str) and len(date_str) == 10 and date_str[2] == '/' and date_str[5] == '/':
        return date_str

    # Essayer de parser avec dateutil (format ISO, etc.)
    try:
        from dateutil import parser
        date_obj = parser.parse(date_str)
        return date_obj.strftime('%d/%m/%Y')  # Format européen
    except ImportError:
        # Fallback: parser manuellement le format ISO
        sys.stderr.write(f"Warning: dateutil not installed, using manual ISO parsing for: {date_str}\n")
        pass  # Continue vers le parsing manuel
    except Exception as e:
        # Log l'erreur mais continue vers le parsing manuel
        sys.stderr.write(f"Warning: dateutil parsing failed for '{date_str}': {e}\n")
        pass

    # Fallback: Parser manuellement le format ISO (YYYY-MM-DDTHH:MM:SS.sssZ)
    try:
        if isinstance(date_str, str):
            # Extraire juste la partie date (YYYY-MM-DD) du format ISO
            if 'T' in date_str:
                date_part = date_str.split('T')[0]
                # Valider que c'est bien au format YYYY-MM-DD
                if len(date_part) == 10 and date_part[4] == '-' and date_part[7] == '-':
                    # Convertir YYYY-MM-DD en DD/MM/YYYY
                    parts = date_part.split('-')
                    return f"{parts[2]}/{parts[1]}/{parts[0]}"

            # Essayer de parser avec datetime standard (sans dateutil)
            from datetime import datetime
            # Essayer plusieurs formats courants
            for fmt in ['%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%SZ', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    return date_obj.strftime('%d/%m/%Y')  # Format européen
                except ValueError:
                    continue
    except Exception as e:
        sys.stderr.write(f"Error: All date parsing methods failed for '{date_str}': {e}\n")

    # Dernier recours: retourner une chaîne vide pour éviter des données incorrectes
    sys.stderr.write(f"Error: Could not parse date '{date_str}', returning empty string\n")
    return ""

def get_gestionnaire_name(gestionnaire):
    """Extracts gestionnaire name from object or returns string directly."""
    if not gestionnaire:
        return ""

    # Si c'est un dictionnaire/objet (cas où Prisma inclut la relation)
    if isinstance(gestionnaire, dict):
        prenom = gestionnaire.get('prenom', '') or ''
        nom = gestionnaire.get('nom', '') or ''
        full_name = f"{prenom} {nom}".strip()
        return full_name if full_name else gestionnaire.get('id', '')

    # Si c'est déjà une chaîne, la retourner
    if isinstance(gestionnaire, str):
        return gestionnaire

    # Autre type inattendu
    return str(gestionnaire) if gestionnaire else ""

def create_excel_export(users_data, output_path):
    """Creates an Excel file from user data."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Usagers"

    headers = [
        "Nom", "Prénom", "Date de naissance", "Genre", "Nationalité",
        "Langue", "Téléphone", "Email",
        "Adresse Complète", "Rue", "Numéro", "Boîte", "Code Postal", "Ville",
        "Secteur", "Statut Séjour", "Date Ouverture", "Date Clôture", "État",
        "Antenne", "Gestionnaire", "Premier Contact", "Notes Générales",
        "Procédure Expulsion?", "Date Réception PrevExp", "Date Requête PrevExp",
        "Date VAD PrevExp", "Décision PrevExp", "Commentaire PrevExp",
        "Type Logement", "Date Entrée Logement", "Date Sortie Logement",
        "Motif Sortie Logement", "Destination Sortie Logement", "Propriétaire Logement",
        "Loyer Logement", "Charges Logement", "Commentaire Logement",
        "Problématiques", "Actions de Suivi"
    ]

    # Style pour les en-têtes
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF") # Police plus grande, toujours en gras
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) # Permettre le retour à la ligne pour les en-têtes longs
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Appliquer les styles aux en-têtes et définir la hauteur de la ligne d'en-tête
    sheet.row_dimensions[1].height = 30 # Augmenter la hauteur de la ligne d'en-tête

    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}1"]
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_alignment

    # Styles pour les cellules de données
    data_font = Font(name='Calibri', size=11) # Police légèrement plus grande pour les données
    data_alignment_default = Alignment(horizontal="left", vertical="top", wrap_text=False)
    data_alignment_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    light_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    wrap_text_columns = ["Problématiques", "Actions de Suivi"]

    for row_idx, user in enumerate(users_data, 0):
        row_num_excel = row_idx + 2

        adresse_data = user.get('adresse') if isinstance(user.get('adresse'), dict) else {}
        logement_details_data = {}
        raw_logement_details = user.get('logementDetails')
        if isinstance(raw_logement_details, dict):
            logement_details_data = raw_logement_details
        elif isinstance(raw_logement_details, str) and raw_logement_details.strip().startswith('{'):
            try:
                logement_details_data = json.loads(raw_logement_details)
                if not isinstance(logement_details_data, dict):
                    logement_details_data = {}
            except json.JSONDecodeError:
                logement_details_data = {"commentaire": raw_logement_details}
        elif isinstance(raw_logement_details, str):
             logement_details_data = {"commentaire": raw_logement_details}

        row_data = [
            user.get("nom", ""),
            user.get("prenom", ""),
            format_date(user.get("dateNaissance")),
            user.get("genre", ""),
            user.get("nationalite", ""),
            user.get("langue", ""),
            user.get("telephone", ""),
            user.get("email", ""),
            f"{adresse_data.get('rue', '')} {adresse_data.get('numero', '')}, {adresse_data.get('codePostal', '')} {adresse_data.get('ville', '')}",
            adresse_data.get("rue", ""),
            adresse_data.get("numero", ""),
            adresse_data.get("boite", ""),
            adresse_data.get("codePostal", ""),
            adresse_data.get("ville", ""),
            user.get("secteur", ""),
            user.get("statutSejour", ""),
            format_date(user.get("dateOuverture")),
            format_date(user.get("dateCloture")),
            user.get("etat", ""),
            user.get("antenne", ""),
            get_gestionnaire_name(user.get("gestionnaire")),
            user.get("premierContact", ""),
            user.get("notesGenerales", ""),
            "Oui" if user.get("hasPrevExp") else "Non",
            format_date(user.get("prevExpDateReception")),
            format_date(user.get("prevExpDateRequete")),
            format_date(user.get("prevExpDateVad")),
            user.get("prevExpDecision", ""),
            user.get("prevExpCommentaire", ""),
            logement_details_data.get("typeLogement", ""),
            format_date(logement_details_data.get("dateEntree")),
            format_date(logement_details_data.get("dateSortie")),
            logement_details_data.get("motifSortie", ""),
            logement_details_data.get("destinationSortie", ""),
            logement_details_data.get("proprietaire", ""),
            logement_details_data.get("loyer", ""),
            logement_details_data.get("charges", ""),
            logement_details_data.get("commentaire", ""),
            ", ".join([f"{p.get('type', '')}: {p.get('description', '')}" for p in user.get("problematiques", []) if p.get('type') or p.get('description')]),
            ", ".join([f"{format_date(a.get('date'))} - {a.get('type', '')}: {a.get('description', '')}" for a in user.get("actions", []) if a.get('date') or a.get('type') or a.get('description')])
        ]

        current_row_fill = light_fill if row_idx % 2 == 0 else no_fill

        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f"{col_letter}{row_num_excel}"]
            cell.value = cell_data
            cell.font = data_font # Appliquer la police de données
            cell.border = thin_border
            cell.fill = current_row_fill

            current_header_title = headers[col_num - 1]
            if current_header_title in wrap_text_columns:
                cell.alignment = data_alignment_wrap
            else:
                cell.alignment = data_alignment_default

    # Ajuster la largeur des colonnes
    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = 0

        # Prendre en compte la longueur de l'en-tête (avec une police potentiellement plus grande)
        # On peut estimer un facteur d'ajustement si la police de l'en-tête est plus grande
        # ou simplement s'assurer que la largeur est suffisante pour l'en-tête.
        header_cell_value = sheet[f"{col_letter}1"].value
        if header_cell_value:
             # Si l'en-tête peut aller à la ligne, on peut prendre la plus longue "partie" ou une estimation
             # Pour simplifier, on prend la longueur totale, wrap_text dans l'en-tête aidera.
             max_length = len(str(header_cell_value)) * 1.1 # Petit facteur pour police plus grande/gras

        column_data_cells = sheet[col_letter][1:] # Cellules de données uniquement
        for cell in column_data_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if header_title in wrap_text_columns:
                        max_length = max(max_length, min(cell_length, 50))
                    elif header_title in ["Notes Générales", "Commentaire Logement", "Commentaire PrevExp", "Adresse Complète"]:
                        max_length = max(max_length, min(cell_length, 70))
                    else:
                        max_length = max(max_length, cell_length)
            except:
                pass

        # Ajuster la largeur, s'assurer qu'elle est au moins suffisante pour l'en-tête
        # La largeur de l'en-tête est déjà dans max_length, on ajoute une marge.
        adjusted_width = (max_length + 3) # Augmenter légèrement la marge
        sheet.column_dimensions[col_letter].width = adjusted_width


    sheet.freeze_panes = 'A2'
    workbook.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.stderr.write("Usage: python export_users_excel.py <input_json_path> <output_excel_path>\n")
        sys.exit(1)

    input_json_path = sys.argv[1]
    output_path = sys.argv[2]

    try:
        with open(input_json_path, 'r') as f:
            users_data = json.load(f)

    except FileNotFoundError:
        sys.stderr.write(f"Error: Input JSON file not found at {input_json_path}\n")
        sys.exit(1)
    except json.JSONDecodeError as e:
        sys.stderr.write(f"Error decoding JSON from {input_json_path}: {e.msg}\n")
        sys.exit(1)
    except Exception as ex:
        sys.stderr.write(f"An unexpected error occurred while reading {input_json_path}: {ex}\n")
        sys.exit(1)

    if not isinstance(users_data, list):
         sys.stderr.write("Error: Expected a list of users from input JSON file\n")
         sys.exit(1)

    create_excel_export(users_data, output_path)
    sys.stdout.write(f"Excel file created successfully at {output_path}\n")
